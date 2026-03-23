import subprocess, sys
subprocess.call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\tanak\Desktop\chemical_ordering_monitoring_system_dotnet\Database_Info_Draft_20032026.xlsx')

with open(r'C:\Users\tanak\Desktop\chemical_ordering_monitoring_system_dotnet\excel_output.txt', 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f'=== Sheet: {sheet_name} ===\n')
        for row in ws.iter_rows(values_only=True):
            f.write(' | '.join([str(c) if c is not None else '' for c in row]) + '\n')

print('Done. Written to excel_output.txt')
